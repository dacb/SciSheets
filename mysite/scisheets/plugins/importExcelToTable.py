"""
Imports an Excel file and adds it to the table
"""

import pandas as pd
import openpyxl
import unicodedata

def _importExcelToDataframe(filepath, worksheet=None):
  """
  Reads the excel file into a dataframe.
  :param str filepath: full path to CSV file
  :param str worksheet: worksheet to import. Default is first.
  :return pandas.DataFrame:
  :raises IOError, ValueError:
  """
  wb = openpyxl.load_workbook(filename=filepath, read_only=True)
  if worksheet is None:
    worksheet = wb.sheetnames[0]
  if not worksheet in wb.sheetnames:
    raise ValueError("Worksheet %s not found in file %s"  \
        % (worksheet, filepath))
  ws = wb[worksheet]
  data = {}
  rows = [ [cell.value for cell in row if cell.value is not None]  \
           for row in ws.rows ]
  names = []
  for name in rows[0]:
    if not (isinstance(name, str) or isinstance(name, unicode)):
      raise ValueError("Invalid column header in %s" % filepath)
    uni = unicode(name)
    name = unicodedata.normalize('NFC', uni).encode('ascii', 'ignore')
    name = name.replace(' ', '')
    names.append(str(name))
    data[name] = []
  data_rows = list(rows)
  del data_rows[0]
  row_num = 1
  num_names = len(names)
  for row in data_rows:
    if len(row) < len(names):
      msg = "Missing data in row %d of imported worksheet" % row_num
      # Issue warningraise ValueError(msg)
    row_num += 1
    num_in_row = len(row)
    for idx in range(num_in_row):
      data[names[idx]].append(row[idx])
    for idx in range(num_in_row, num_names):
      data[names[idx]].append(None)
  df = pd.DataFrame(data)
  return df

def importExcelToTable(s, 
                       filepath, 
                       worksheet=None, 
                       names=None,
                       column_position=None):
  """
  Imports the specified columns from the excel file.
  Columns that don't exist in the current table are created.
  :param API s: API object
  :param str filepath: full path to CSV file
  :param str worksheet: worksheet to import. Default is first.
  :param list-of-str names: names of columns in worksheet to add in
                            the order they are to be added
  :param str column_position: name of the column to place after
  :return list-of-str: column names imported
  """
  df = _importExcelToDataframe(filepath, worksheet=worksheet)
  imported_names = s.addColumnsToTableFromDataframe(df, names=names, column_position=column_position)
  return imported_names
